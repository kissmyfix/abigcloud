#!/usr/bin/env python3
"""Verification of tn-idb-debt-master-2021-2023.csv.

Pass 1 - Coverage: three years present, row counts recounted independently
         from the source workbook, no row missing an entity.
Pass 2 - Field integrity: money columns numeric, no column shift, 'No Debt'
         rows carry no amounts, debt rows carry an amount.
Pass 3 - Anchors: figures read by eye off the source reproduce.

Exits nonzero if any check fails.
"""

import os
import sys

import pandas as pd
import openpyxl

BASE = os.path.join(
    os.path.dirname(os.path.dirname(os.path.dirname(os.path.abspath(__file__)))),
    "state_of_tennessee", "tn_comptroller_pilot_reports")
SRC = os.path.join(BASE, "tn_comptroller_archived", "idb_debt-reports.xlsx")
MASTER = os.path.join(BASE, "derived", "tn-idb-debt-master-2021-2023.csv")

FAILURES = []


def check(ok, label, detail=""):
    print(f"  [{'PASS' if ok else 'FAIL'}] {label}{(' - ' + detail) if detail else ''}")
    if not ok:
        FAILURES.append(label)


def pass1(df):
    print("\nPASS 1 - Coverage")
    years = sorted(df["YEAR"].unique())
    check(years == [2021, 2022, 2023], "years 2021-2023 present", f"got {years}")
    check(df["ENTITY"].isna().sum() == 0, "no row missing an entity",
          f"{df['ENTITY'].isna().sum()} missing")
    wb = openpyxl.load_workbook(SRC, read_only=True, data_only=True)
    for sheet in wb.sheetnames:
        n = sum(1 for i, r in enumerate(wb[sheet].iter_rows(values_only=True))
                if i > 0 and any(str(c).strip() for c in r if c is not None))
        got = int((df["YEAR"] == int(sheet)).sum())
        check(got == n, f"{sheet} row count reconciles with source",
              f"parsed {got}, source {n}")


def pass2(df):
    print("\nPASS 2 - Field integrity")
    for col in ("ORIGINAL_AMOUNT", "OUTSTANDING_FYE"):
        vals = pd.to_numeric(df[col], errors="coerce")
        check(vals.notna().sum() == df[col].notna().sum(), f"{col} fully numeric",
              f"{df[col].notna().sum() - vals.notna().sum()} unparseable")
        neg = int((vals < 0).sum())
        check(neg == 0, f"{col} has no negative values", f"{neg} negative")

    types = set(df["DEBT_TYPE"].dropna().unique())
    check(types <= {"Conduit", "Direct", "Non-Debt", "No Debt"},
          "DEBT_TYPE holds only known categories", f"got {sorted(types)}")

    nodebt = df[~df["HAS_DEBT"]]
    stray = nodebt[nodebt["ORIGINAL_AMOUNT"].notna()
                   | nodebt["OUTSTANDING_FYE"].notna()]
    check(len(stray) == 0, "'No Debt' rows carry no amounts",
          f"{len(stray)} row(s) do")

    debt = df[df["HAS_DEBT"]]
    missing = int(debt["ORIGINAL_AMOUNT"].isna().sum())
    check(missing == 0, "every debt row states an original amount",
          f"{missing} row(s) do not")

    # Outstanding-at-FYE is widely left blank by filers; report, do not fail.
    blank = int(debt["OUTSTANDING_FYE"].isna().sum())
    print(f"  [INFO] {blank} of {len(debt)} debt rows leave OUTSTANDING_FYE blank "
          f"({blank / len(debt):.0%}) - a filer omission, present in the source")


def pass3(df):
    print("\nPASS 3 - Anchors")
    gall = df[df["ENTITY"].str.contains("Gallatin", na=False)]
    check(len(gall) == 3, "Gallatin IDB files once per year", f"{len(gall)} rows")
    check(not gall["HAS_DEBT"].any(),
          "Gallatin IDB reports No Debt in all three years",
          f"{int(gall['HAS_DEBT'].sum())} year(s) report debt")

    sumner = df[df["ENTITY"].str.contains("County of Sumner", na=False)]
    check(len(sumner) == 3, "the separate Sumner County IDB files once per year",
          f"{len(sumner)} rows")
    check(not sumner["HAS_DEBT"].any(),
          "Sumner County IDB reports No Debt in all three years",
          f"{int(sumner['HAS_DEBT'].sum())} year(s) report debt")

    port = df[df["ENTITY"].str.contains("Portland", na=False) & df["HAS_DEBT"]]
    for label, pat, amt in (("North American Stamping", "North American Stamping", 28_000_000),
                            ("Shoals/Solon", "Shoals", 25_500_000),
                            ("SIF Portland/RB Distribution", "SIF Portland", 50_000_000)):
        rows = port[port["DEBT_NAME"].str.contains(pat, na=False, case=False)]
        hit = (pd.to_numeric(rows["ORIGINAL_AMOUNT"], errors="coerce") == amt).any()
        check(bool(hit), f"Portland IDB conduit debt for {label} is ${amt:,}",
              f"values seen: {sorted(set(rows['ORIGINAL_AMOUNT'].dropna()))}")

    # No data center anywhere in the state carries IDB debt.
    # \bmeta\b, not bare 'meta' - otherwise Taco Metals and Bridgestone
    # Metalpha (steel and tire manufacturing) match as substrings.
    pat = (r"data ?cent|\bmeta\b|woolhawk|archer|hyperscale|facebook|google|"
           r"amazon|microsoft|oracle|vantage")
    blob = (df["ENTITY"].fillna("") + " " + df["DEBT_NAME"].fillna("") + " "
            + df["PROJECT"].fillna(""))
    hits = df[blob.str.contains(pat, case=False, regex=True)]
    check(len(hits) == 0, "no data-center project carries IDB debt statewide",
          f"{len(hits)} hit(s): {sorted(set(hits['DEBT_NAME'].dropna()))[:3]}")


def main():
    df = pd.read_csv(MASTER)
    print(f"Loaded {len(df):,} rows from {os.path.basename(MASTER)}")
    pass1(df)
    pass2(df)
    pass3(df)
    print("\n" + "=" * 60)
    if FAILURES:
        print(f"{len(FAILURES)} CHECK(S) FAILED:")
        for f in FAILURES:
            print(f"  - {f}")
    else:
        print("ALL CHECKS PASSED")
    sys.exit(1 if FAILURES else 0)


if __name__ == "__main__":
    main()

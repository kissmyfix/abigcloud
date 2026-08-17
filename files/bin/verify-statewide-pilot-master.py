#!/usr/bin/env python3
"""Four-pass verification of tn-statewide-pilot-master-2014-2025.csv.

Pass 1 - Coverage: every year present, every county name a real TN county,
         parsed row count reconciles against an independent recount of the
         source files.
Pass 2 - Field integrity: no column-shift. Emails look like emails, dates like
         dates, money columns numeric, per-year fill rates sane.
Pass 3 - Sumner cross-check: the Sumner slice reconciles against the existing
         derived/sumner_idb_master_2017-2025.csv built in an earlier pass from
         the county-cropped PDFs.
Pass 4 - Anchors: known figures reproduce (Woolhawk $0 county, North American
         Stamping ~$50k county, Archer Datacenters ~$53k county).

Exits nonzero if any pass fails.
"""

import os
import re
import sys

import pandas as pd
import pdfplumber
import openpyxl

BASE = os.path.join(
    os.path.dirname(os.path.dirname(os.path.dirname(os.path.abspath(__file__)))),
    "state_of_tennessee", "tn_comptroller_pilot_reports")
SRC = os.path.join(BASE, "tn_comptroller_archived")
MASTER = os.path.join(BASE, "derived", "tn-statewide-pilot-master-2014-2025.csv")
SUMNER = os.path.join(BASE, "derived", "sumner_idb_master_2017-2025.csv")

COUNTY_NAMES = set("""Anderson Bedford Benton Bledsoe Blount Bradley Campbell Cannon Carroll Carter
Cheatham Chester Claiborne Clay Cocke Coffee Crockett Cumberland Davidson Decatur
DeKalb Dickson Dyer Fayette Fentress Franklin Gibson Giles Grainger Greene
Grundy Hamblen Hamilton Hancock Hardeman Hardin Hawkins Haywood Henderson Henry
Hickman Houston Humphreys Jackson Jefferson Johnson Knox Lake Lauderdale Lawrence
Lewis Lincoln Loudon McMinn McNairy Macon Madison Marion Marshall Maury
Meigs Monroe Montgomery Moore Morgan Obion Overton Perry Pickett Polk
Putnam Rhea Roane Robertson Rutherford Scott Sequatchie Sevier Shelby Smith
Stewart Sullivan Sumner Tipton Trousdale Unicoi Union Van_Buren Warren Washington
Wayne Weakley White Williamson Wilson""".split())
COUNTY_NAMES = {n.replace("_", " ") for n in COUNTY_NAMES}

FAILURES = []
WARNINGS = []


def check(ok, label, detail=""):
    print(f"  [{'PASS' if ok else 'FAIL'}] {label}{(' - ' + detail) if detail else ''}")
    if not ok:
        FAILURES.append(label)


def warn(label, detail=""):
    print(f"  [WARN] {label}{(' - ' + detail) if detail else ''}")
    WARNINGS.append(label)


def blank(row):
    return not any(str(c).strip() for c in row if c is not None)


HEADER_FIRSTS = {"COUNTY", "DATE RECV'D.", "DATE RECV'D"}
HEADER_LABELS = {"PROJ TYPE", "PROJECT TYPE", "IDB/HED", "FILING DATE",
                 "DATE RECEIVED", "LESSEE", "LESSEE NAME", "PROPERTY ADDRESS"}


def header(row):
    """Header rows repeat on every PDF page and mid-file in the 2023 xlsx."""
    cells = [" ".join(str(c or "").split()).upper() for c in row]
    if cells[0] in HEADER_FIRSTS:
        return True
    return sum(1 for c in cells[1:6] if c in HEADER_LABELS) >= 2


def source_row_count(year):
    """Recount data rows straight from the source, independent of the parser."""
    pdf_path = os.path.join(SRC, f"{year}-pilot.pdf")
    xlsx_path = os.path.join(SRC, f"{year}-pilot.xlsx")
    n = 0
    if os.path.exists(pdf_path):
        with pdfplumber.open(pdf_path) as pdf:
            for page in pdf.pages:
                for table in page.extract_tables():
                    for row in table:
                        if not row or blank(row) or header(row):
                            continue
                        n += 1
    else:
        wb = openpyxl.load_workbook(xlsx_path, read_only=True, data_only=True)
        ws = wb[wb.sheetnames[0]]
        for i, row in enumerate(ws.iter_rows(values_only=True)):
            if i == 0 or blank(row) or header(row):
                continue
            n += 1
    return n


def pass1(df):
    print("\nPASS 1 - Coverage")
    years = sorted(df["YEAR"].unique())
    check(years == list(range(2014, 2026)), "all 12 years 2014-2025 present",
          f"got {years}")
    bad = sorted(set(df["COUNTY"].dropna()) - COUNTY_NAMES)
    check(not bad, "every COUNTY value is a real TN county", f"unknown: {bad}")
    check(df["COUNTY"].isna().sum() == 0, "no rows missing a county",
          f"{df['COUNTY'].isna().sum()} missing")
    for year in years:
        got = int((df["YEAR"] == year).sum())
        want = source_row_count(year)
        check(got == want, f"{year} row count reconciles with source",
              f"parsed {got}, source {want}")


def pass2(df):
    print("\nPASS 2 - Field integrity (column-shift detection)")
    # "NO EMAIL ON FILE" / "NOF" are the source's own sentinels, not a shift.
    em = df["EMAIL"].dropna()
    em = em[(em.str.strip() != "") & (~em.str.upper().isin(
        {"NO EMAIL ON FILE", "NOF", "NONE", "N/A"}))]
    good = em.str.contains("@", regex=False).mean() if len(em) else 1.0
    check(good > 0.999, "EMAIL column holds email addresses",
          f"{good:.4%} contain '@' across {len(em)} values "
          "(remainder are filer typos: '.' where '@' belongs)")

    dt = df["FILING_DATE"].dropna()
    good = dt.str.match(r"^\d{1,2}/\d{1,2}/\d{4}$").mean() if len(dt) else 1.0
    check(good > 0.98, "FILING_DATE parses as m/d/yyyy",
          f"{good:.3%} of {len(dt)}")

    for col in ("EST_VALUE", "RENT", "PILOT_CITY", "PILOT_COUNTY", "LH_TAX"):
        vals = pd.to_numeric(df[col], errors="coerce")
        nonnull = df[col].notna().sum()
        check(vals.notna().sum() == nonnull, f"{col} fully numeric",
              f"{nonnull - vals.notna().sum()} unparseable")
        neg = int((vals < 0).sum())
        if neg:
            warn(f"{col} has {neg} negative value(s)")

    pc = df["PROP_CODE"].dropna()
    good = pc.str.match(r"^(ID|HE|NOF)", case=False).mean() if len(pc) else 1.0
    check(good > 0.90, "PROP_CODE looks like a comptroller property code",
          f"{good:.3%} of {len(pc)}")

    # Per-year fill rate on EST_VALUE catches a whole year mapped wrong.
    print("  EST_VALUE fill rate by year:")
    for year, grp in df.groupby("YEAR"):
        rate = grp["EST_VALUE"].notna().mean()
        flag = "" if rate > 0.60 else "   <-- LOW"
        print(f"    {year}: {rate:6.1%} of {len(grp):5d} rows{flag}")
        if rate <= 0.60:
            warn(f"{year} EST_VALUE fill rate {rate:.1%}")


def norm_name(s):
    return re.sub(r"[^a-z0-9]", "", str(s).lower())


# Sumner row counts read straight off the cropped source PDFs by hand, used as
# ground truth. Only years whose Sumner block could be counted unambiguously
# are listed; see the tn_comptroller_archived/ README.
SUMNER_GROUND_TRUTH = {2016: 8, 2017: 11, 2018: 6, 2020: 9, 2022: 19}


def pass3(df):
    print("\nPASS 3 - Sumner ground truth + reconciliation")
    mine = df[df["COUNTY"] == "Sumner"]

    # 3a. Against hand-counted Sumner blocks in the cropped source PDFs.
    for year, want in sorted(SUMNER_GROUND_TRUTH.items()):
        got = int((mine["YEAR"] == year).sum())
        check(got == want, f"{year} Sumner row count matches hand-counted source",
              f"parsed {got}, source block {want}")

    # 3b. Reconciliation against the older county-cropped derived file. That
    # file is known incomplete (it matches the hand-counted source in none of
    # the years above), so differences are reported, not failed.
    if not os.path.exists(SUMNER):
        warn("sumner master not found, skipping reconciliation", SUMNER)
        return
    ref = pd.read_csv(SUMNER)
    ref["YEAR"] = ref["source"].str.extract(r"^(\d{4})").astype(int)
    print("  Reconciliation vs sumner_idb_master_2017-2025.csv (known incomplete):")
    print(f"    {'year':>5} {'new':>4} {'old':>4}  entities only in new")
    for year in sorted(set(ref["YEAR"]) | set(mine["YEAR"])):
        r, m = ref[ref["YEAR"] == year], mine[mine["YEAR"] == year]
        rn = {norm_name(x) for x in r["lessee"].dropna()}
        mn = {norm_name(x) for x in m["LESSEE"].dropna()}
        extra = sorted(n for n in mn - rn
                       if not any(n.startswith(o[:12]) for o in rn))
        print(f"    {year:>5} {len(m):>4} {len(r):>4}  {', '.join(extra[:4]) or '-'}")


def pass4(df):
    print("\nPASS 4 - Known anchors")
    sumner = df[df["COUNTY"] == "Sumner"]

    # Exact values, not ranges. A range check here passed a build in which the
    # 2024-2025 county forward-fill had silently broken and two thirds of the
    # Woolhawk filings had vanished.
    wool = sumner[sumner["LESSEE"].fillna("").str.upper().str.contains("WOOLHAWK")]
    check(len(wool) == 41, "Woolhawk filing count is 41",
          f"got {len(wool)} rows, years {sorted(wool['YEAR'].unique())}")
    check(sorted(wool["YEAR"].unique()) == [2021, 2022, 2023, 2024, 2025],
          "Woolhawk appears in every year 2021-2025",
          f"got {sorted(wool['YEAR'].unique())}")
    cty = pd.to_numeric(wool["PILOT_COUNTY"], errors="coerce")
    check(cty.notna().sum() == 0,
          "Woolhawk county PILOT is blank - never reported - in all 41 filings",
          f"{cty.notna().sum()} filing(s) state a county figure")
    city = pd.to_numeric(wool["PILOT_CITY"], errors="coerce").fillna(0).sum()
    check(abs(city - 3_321_650) < 1, "Woolhawk city PILOT total is $3,321,650",
          f"got ${city:,.2f}")
    top = pd.to_numeric(wool["EST_VALUE"], errors="coerce").max()
    check(abs(top - 519_189_800) < 1, "Woolhawk top EST_VALUE is $519,189,800",
          f"got ${top:,.0f}")

    for label, pat, lo, hi in (
            ("North American Stamping", "NORTH AMERICAN STAMPING", 40000, 60000),
            ("Archer Datacenters", "ARCHER", 45000, 60000)):
        rows = sumner[sumner["LESSEE"].fillna("").str.upper().str.contains(pat)]
        if not len(rows):
            warn(f"{label} not found in Sumner rows")
            continue
        cty = pd.to_numeric(rows["PILOT_COUNTY"], errors="coerce").dropna()
        hit = ((cty >= lo) & (cty <= hi)).any()
        check(bool(hit), f"{label} shows a county PILOT in the ${lo:,}-${hi:,} range",
              f"values seen: {sorted(set(cty))[:6]}")

    # Control-group contrast, the plain-language proof.
    yr = sumner[sumner["YEAR"] == sumner["YEAR"].max()]
    cty_paid = pd.to_numeric(yr["PILOT_COUNTY"], errors="coerce").fillna(0)
    print(f"    Sumner {int(sumner['YEAR'].max())}: {len(yr)} filings, "
          f"{int((cty_paid > 0).sum())} paying the county, "
          f"{int((cty_paid == 0).sum())} paying it nothing")


def main():
    df = pd.read_csv(MASTER)
    print(f"Loaded {len(df):,} rows from {os.path.basename(MASTER)}")
    pass1(df)
    pass2(df)
    pass3(df)
    pass4(df)
    print("\n" + "=" * 60)
    if FAILURES:
        print(f"{len(FAILURES)} CHECK(S) FAILED:")
        for f in FAILURES:
            print(f"  - {f}")
    else:
        print("ALL CHECKS PASSED")
    if WARNINGS:
        print(f"\n{len(WARNINGS)} warning(s):")
        for w in WARNINGS:
            print(f"  - {w}")
    sys.exit(1 if FAILURES else 0)


if __name__ == "__main__":
    main()

#!/usr/bin/env python3
"""Build the IDB debt master CSV from the Comptroller's IDB debt report workbook.

Source: state_of_tennessee/tn_comptroller_pilot_reports/tn_comptroller_archived/
        idb_debt-reports.xlsx (sheets 2021, 2022, 2023)
Output: state_of_tennessee/tn_comptroller_pilot_reports/derived/
        tn-idb-debt-master-2021-2023.csv

Separate schema from the PILOT master - this is debt reporting, not abatement
reporting, and the two share no columns beyond the year.

Needs pandas + openpyxl.
"""

import os

import pandas as pd
import openpyxl

BASE = os.path.join(
    os.path.dirname(os.path.dirname(os.path.dirname(os.path.abspath(__file__)))),
    "state_of_tennessee", "tn_comptroller_pilot_reports")
SRC = os.path.join(BASE, "tn_comptroller_archived", "idb_debt-reports.xlsx")
OUT = os.path.join(BASE, "derived", "tn-idb-debt-master-2021-2023.csv")

OUT_COLS = ["YEAR", "ENTITY", "DEBT_NAME", "ORIGINAL_AMOUNT", "OUTSTANDING_FYE",
            "PROJECT", "DEBT_TYPE", "HAS_DEBT", "SOURCE_SHEET"]

# Source column order, left to right.
FIELDS = ["ENTITY", "DEBT_NAME", "ORIGINAL_AMOUNT", "OUTSTANDING_FYE",
          "PROJECT", "DEBT_TYPE"]
NUMERIC_FIELDS = {"ORIGINAL_AMOUNT", "OUTSTANDING_FYE"}


def clean_str(v):
    if v is None:
        return None
    v = " ".join(str(v).replace("\xad", "").split())
    return v if v and v.lower() != "nan" else None


def clean_num(v):
    s = clean_str(v)
    if s is None:
        return None
    s = s.replace(",", "").replace("$", "")
    if s.startswith("(") and s.endswith(")"):
        s = "-" + s[1:-1]
    try:
        return float(s)
    except ValueError:
        return None


def main():
    wb = openpyxl.load_workbook(SRC, read_only=True, data_only=True)
    rows = []
    for sheet in wb.sheetnames:
        year = int(sheet)
        current_entity = None
        n = 0
        for i, raw in enumerate(wb[sheet].iter_rows(values_only=True)):
            raw = list(raw)[:len(FIELDS)]
            if i == 0 or not any(clean_str(c) for c in raw):
                continue
            rec = {c: None for c in OUT_COLS}
            rec["YEAR"] = year
            rec["SOURCE_SHEET"] = sheet
            for name, val in zip(FIELDS, raw):
                rec[name] = clean_num(val) if name in NUMERIC_FIELDS else clean_str(val)
            # Entity is stated once and carries down its debt rows.
            if rec["ENTITY"]:
                current_entity = rec["ENTITY"]
            else:
                rec["ENTITY"] = current_entity
            rec["HAS_DEBT"] = (rec["DEBT_TYPE"] or "").strip().lower() != "no debt"
            rows.append(rec)
            n += 1
        print(f"{sheet}: {n} rows")

    df = pd.DataFrame(rows, columns=OUT_COLS)
    df = df.sort_values(["YEAR", "ENTITY", "DEBT_NAME"], kind="stable",
                        na_position="last").reset_index(drop=True)
    df.to_csv(OUT, index=False)
    print(f"\nWrote {len(df)} rows to {OUT}")
    print(f"{df['ENTITY'].nunique()} distinct entities across the three years")
    print(df.groupby("YEAR")["HAS_DEBT"].agg(
        rows="size", with_debt="sum").to_string())


if __name__ == "__main__":
    main()

#!/usr/bin/env python3
"""Build the statewide TN PILOT master CSV from the Comptroller's annual reports.

Source: state_of_tennessee/tn_comptroller_pilot_reports/tn_comptroller_archived/
        2014-2022 PDFs, 2023-2025 XLSX
Output: state_of_tennessee/tn_comptroller_pilot_reports/derived/
        tn-statewide-pilot-master-2014-2025.csv

Schema is the canonical one established by build-hamilton-master.py, with two
columns appended: PROP_CLASS (the A/B property class carried only by the
2014-2015 layout) and SOURCE_FILE.

Needs pdfplumber + pandas + openpyxl. Build a venv if the imports fail:
    python3 -m venv venv && venv/bin/pip install pdfplumber pandas openpyxl
"""

import glob
import os

import pandas as pd
import pdfplumber
import openpyxl

BASE = os.path.join(
    os.path.dirname(os.path.dirname(os.path.dirname(os.path.abspath(__file__)))),
    "state_of_tennessee", "tn_comptroller_pilot_reports")
SRC = os.path.join(BASE, "tn_comptroller_archived")
OUT = os.path.join(BASE, "derived", "tn-statewide-pilot-master-2014-2025.csv")

OUT_COLS = ["YEAR", "COUNTY", "PROJ_TYPE", "FILING_DATE", "CASE_NO", "LESSEE",
            "PROPERTY_ADDRESS", "CITY", "PARCEL_ID", "PROP_TYPE", "PROP_CODE",
            "CONTACT", "CONTACT_TITLE", "EMAIL", "EST_VALUE", "RENT",
            "PILOT_CITY", "PILOT_COUNTY", "LH_TAX", "LEASE_BEGIN", "LEASE_END",
            "REPORTING_FLAG", "SOURCE_NOTE", "PROP_CLASS", "SOURCE_FILE"]

NUMERIC_FIELDS = {"EST_VALUE", "RENT", "PILOT_CITY", "PILOT_COUNTY", "LH_TAX"}

# TN county codes 1-95. 73 of these pairs are proven directly by the 2016 and
# 2017 reports, which print name and code side by side and agree on every
# shared county. Codes 29/48/64/81 (Grainger, Lake, Moore, Stewart) appear in
# the 2014-2015 code-only layout but in no name-bearing report; they are filled
# from the published alphabetical list and every row using them is flagged in
# SOURCE_NOTE. See the tn_comptroller_archived/ README.
COUNTY_NAMES = """Anderson Bedford Benton Bledsoe Blount Bradley Campbell Cannon Carroll Carter
Cheatham Chester Claiborne Clay Cocke Coffee Crockett Cumberland Davidson Decatur
DeKalb Dickson Dyer Fayette Fentress Franklin Gibson Giles Grainger Greene
Grundy Hamblen Hamilton Hancock Hardeman Hardin Hawkins Haywood Henderson Henry
Hickman Houston Humphreys Jackson Jefferson Johnson Knox Lake Lauderdale Lawrence
Lewis Lincoln Loudon McMinn McNairy Macon Madison Marion Marshall Maury
Meigs Monroe Montgomery Moore Morgan Obion Overton Perry Pickett Polk
Putnam Rhea Roane Robertson Rutherford Scott Sequatchie Sevier Shelby Smith
Stewart Sullivan Sumner Tipton Trousdale Unicoi Union Van_Buren Warren Washington
Wayne Weakley White Williamson Wilson""".split()
COUNTY_BY_CODE = {i + 1: n.replace("_", " ") for i, n in enumerate(COUNTY_NAMES)}
INFERRED_CODES = {29, 48, 64, 81}

# Field order of each layout, index 0 = leftmost column. None = drop.
FORMATS = {
    2014: ["FILING_DATE", "COUNTY_CODE", "PROJ_TYPE", "LESSEE", "CONTACT",
           "PARCEL_ID", "PROP_CODE", "REPORTING_FLAG", "PROP_CLASS",
           "EST_VALUE", "RENT", "PILOT_CITY", "PILOT_COUNTY", "LH_TAX",
           "LEASE_END"],
    2016: ["COUNTY", "COUNTY_CODE", "FILING_DATE", "PROJ_TYPE", "LESSEE",
           "PROPERTY_ADDRESS", "CITY", "CONTACT", "EMAIL", "PARCEL_ID",
           "PROP_CODE", "REPORTING_FLAG", "PROP_CLASS", "EST_VALUE", "RENT",
           "PILOT_CITY", "PILOT_COUNTY", "LH_TAX", "LEASE_BEGIN", "LEASE_END"],
    2017: ["COUNTY", "COUNTY_CODE", "PROJ_TYPE", "FILING_DATE", "CASE_NO",
           "LESSEE", "PROPERTY_ADDRESS", "CITY", "PARCEL_ID", "PROP_CODE",
           "CONTACT", "CONTACT_TITLE", "EMAIL", "EST_VALUE", "RENT",
           "PILOT_CITY", "PILOT_COUNTY", "LH_TAX", "LEASE_BEGIN", "LEASE_END"],
    2018: ["COUNTY", "PROJ_TYPE", "FILING_DATE", "CASE_NO", "LESSEE",
           "PROPERTY_ADDRESS", "CITY", "PARCEL_ID", "PROP_CODE", "CONTACT",
           "CONTACT_TITLE", "EMAIL", "EST_VALUE", "RENT", "PILOT_CITY",
           "PILOT_COUNTY", "LH_TAX", "LEASE_BEGIN", "LEASE_END"],
    2020: ["COUNTY", "PROJ_TYPE", "FILING_DATE", "CASE_NO", "LESSEE",
           "PROPERTY_ADDRESS", "CITY", "PARCEL_ID", "PROP_TYPE", "PROP_CODE",
           "CONTACT", "CONTACT_TITLE", "EMAIL", "EST_VALUE", "RENT",
           "PILOT_CITY", "PILOT_COUNTY", "LH_TAX", "LEASE_BEGIN", "LEASE_END"],
    2021: ["COUNTY", "PROJ_TYPE", "FILING_DATE", "LESSEE", "PROPERTY_ADDRESS",
           "CITY", "PARCEL_ID", "PROP_TYPE", "PROP_CODE", "CONTACT",
           "CONTACT_TITLE", "EMAIL", "EST_VALUE", "RENT", "PILOT_CITY",
           "PILOT_COUNTY", "LH_TAX", "LEASE_BEGIN", "LEASE_END"],
}
FORMATS[2015] = FORMATS[2014]
FORMATS[2019] = FORMATS[2018]
FORMATS[2022] = FORMATS[2021]
# 2023 xlsx carries a proper header row; 2024/2025 are the same field order.
FORMATS[2023] = FORMATS[2024] = FORMATS[2025] = FORMATS[2021]

CODE_ONLY_YEARS = {2014, 2015}


def clean_num(v):
    if v is None:
        return None
    v = str(v).strip().replace(",", "").replace("$", "").replace("\xad", "")
    if v.startswith("(") and v.endswith(")"):
        v = "-" + v[1:-1]
    if v == "" or v.lower() == "nan":
        return None
    try:
        return float(v)
    except ValueError:
        return None


def clean_str(v):
    if v is None:
        return None
    if isinstance(v, float) and pd.isna(v):
        return None
    v = " ".join(str(v).replace("\xad", "").split())
    return v if v and v.lower() != "nan" else None


def clean_date(v):
    """Normalize the xlsx datetime strings; leave PDF dates as filed."""
    s = clean_str(v)
    if s and len(s) >= 10 and s[4] == "-" and s[7] == "-":
        y, m, d = s[:10].split("-")
        return f"{int(m)}/{int(d)}/{y}"
    return s


ADDRESS_HINTS = (" ROAD", " RD", " STREET", " ST", " DRIVE", " DR", " AVENUE",
                 " AVE", " LANE", " LN", " HIGHWAY", " HWY", " PARKWAY", " PKWY",
                 " BLVD", " BOULEVARD", " COURT", " CT", " PIKE", " WAY", " CIRCLE")


def looks_like_address(v):
    if not v:
        return False
    u = " " + v.upper()
    return any(h + " " in u + " " for h in ADDRESS_HINTS)


def map_row(year, field_names, raw_row, source_file):
    rec = {c: None for c in OUT_COLS}
    rec["YEAR"] = year
    rec["SOURCE_FILE"] = source_file
    notes = []
    code = None
    for name, val in zip(field_names, raw_row):
        if name == "COUNTY_CODE":
            code = clean_num(val)
            continue
        if name not in rec:
            continue
        if name in NUMERIC_FIELDS:
            rec[name] = clean_num(val)
        elif name == "FILING_DATE":
            rec[name] = clean_date(val)
        else:
            rec[name] = clean_str(val)

    if year in CODE_ONLY_YEARS:
        if code is None:
            notes.append("county code missing in source row")
        else:
            c = int(code)
            rec["COUNTY"] = COUNTY_BY_CODE.get(c)
            if c in INFERRED_CODES:
                notes.append(
                    f"county name for code {c} inferred from the published "
                    "alphabetical list, not proven by a name-bearing report")
            elif c not in COUNTY_BY_CODE:
                notes.append(f"unrecognized county code {c}")
        # This layout has one combined 'Prop.Desc.' field holding either a
        # parcel ID or a street address.
        if looks_like_address(rec["PARCEL_ID"]):
            rec["PROPERTY_ADDRESS"] = rec["PARCEL_ID"]
            rec["PARCEL_ID"] = None
            notes.append(
                "2014-2015 layout: combined Prop.Desc. field routed to "
                "PROPERTY_ADDRESS - verify against source")

    rec["SOURCE_NOTE"] = "; ".join(notes) if notes else None
    return rec


def is_blank(row):
    return not any(clean_str(c) for c in row)


# Header rows repeat throughout the sources: on every PDF page, and scattered
# mid-file in the 2023 spreadsheet. Detect them by their labels, not position.
HEADER_FIRSTS = {"COUNTY", "DATE RECV'D.", "DATE RECV'D"}
HEADER_LABELS = {"PROJ TYPE", "PROJECT TYPE", "IDB/HED", "FILING DATE",
                 "DATE RECEIVED", "LESSEE", "LESSEE NAME", "PROPERTY ADDRESS"}


def is_header(row):
    if (clean_str(row[0]) or "").upper() in HEADER_FIRSTS:
        return True
    labels = sum(1 for c in row[1:6] if (clean_str(c) or "").upper() in HEADER_LABELS)
    return labels >= 2


def rows_from_pdf(path, year):
    field_names = FORMATS[year]
    fname = os.path.basename(path)
    county_idx = 0 if field_names[0] == "COUNTY" else None
    out, current_county = [], None
    with pdfplumber.open(path) as pdf:
        for page in pdf.pages:
            for table in page.extract_tables():
                for row in table:
                    if not row or is_blank(row) or is_header(row):
                        continue
                    first = clean_str(row[0]) or ""
                    if county_idx is not None:
                        if first:
                            current_county = first
                        row = list(row)
                        row[0] = current_county
                    out.append(map_row(year, field_names, row, fname))
    return out


def rows_from_xlsx(path, year):
    field_names = FORMATS[year]
    fname = os.path.basename(path)
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    # 2025 leaves several working sheets in the workbook; the filed report is
    # the first sheet in every year.
    ws = wb[wb.sheetnames[0]]
    out, current_county = [], None
    for i, row in enumerate(ws.iter_rows(values_only=True)):
        row = list(row)
        if is_blank(row):
            continue
        first = clean_str(row[0]) or ""
        if is_header(row):
            # These headers repeat mid-file. 2023's carries the literal COUNTY
            # label; 2024/2025's carry a county NAME in that cell instead, and
            # that name is the county the following rows belong to - so the
            # row must be skipped as data but still advance the county.
            if first and first.upper() != "COUNTY":
                current_county = first
            continue
        if first:
            current_county = first
        row[0] = current_county
        out.append(map_row(year, field_names, row, fname))
    return out


def main():
    all_rows = []
    for path in sorted(glob.glob(os.path.join(SRC, "*-pilot.pdf"))):
        year = int(os.path.basename(path)[:4])
        recs = rows_from_pdf(path, year)
        print(f"{os.path.basename(path)}: {len(recs)} rows")
        all_rows.extend(recs)
    for path in sorted(glob.glob(os.path.join(SRC, "*-pilot.xlsx"))):
        year = int(os.path.basename(path)[:4])
        recs = rows_from_xlsx(path, year)
        print(f"{os.path.basename(path)}: {len(recs)} rows")
        all_rows.extend(recs)

    df = pd.DataFrame(all_rows, columns=OUT_COLS)
    df = df.sort_values(["YEAR", "COUNTY", "LESSEE"], kind="stable",
                        na_position="last").reset_index(drop=True)
    df.to_csv(OUT, index=False)
    print(f"\nWrote {len(df)} rows to {OUT}")
    print(df.groupby("YEAR").size().to_string())
    flagged = df[df["SOURCE_NOTE"].notna()]
    print(f"\n{len(flagged)} row(s) carry a SOURCE_NOTE")


if __name__ == "__main__":
    main()
